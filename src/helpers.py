# helpers.py
# Utilidades puras y de acceso a datos (sin UI específica de Flet)

import os
import csv

# -------------------------
# Normalización y parsing
# -------------------------
def normalize_string(s: str) -> str:
    return (s or "").strip()

def to_int_safe(v) -> int:
    try:
        s = str(v).strip()
        if s == "" or s.lower() in ("none", "nan", "null"):
            return 0
        return int(float(s.replace(",", "")))
    except Exception:
        return 0

def norm_text(s: str) -> str:
    s = (s or "").lower()
    for a, b in [("á", "a"), ("é", "e"), ("í", "i"), ("ó", "o"), ("ú", "u")]:
        s = s.replace(a, b)
    return s

def _map_headers(headers: list[str]):
    def _norm(h: str):
        h = (h or "").strip().lower()
        for a, b in [("á", "a"), ("é", "e"), ("í", "i"), ("ó", "o"), ("ú", "u")]:
            h = h.replace(a, b)
        return h

    norm = [_norm(h) for h in headers]

    def find(*c):
        for k in c:
            if k in norm:
                return norm.index(k)
        return -1

    i_code = find("codigo", "código", "code", "sku", "clave")
    i_name = find("nombre", "name", "producto")
    i_desc = find("descripcion", "descripción", "description", "detalle")
    i_qty = find("existencia", "existencias", "qty", "cantidad", "stock", "inventario", "exist")

    if i_code < 0 or i_name < 0:
        raise ValueError("Encabezados requeridos: Código y Nombre (Descripción y Existencias opcionales).")

    return i_code, i_name, i_desc, i_qty


def parse_products_from_file(file_meta) -> list[dict]:
    """
    Retorna una lista de dicts: {"code","name","description","qty"}
    Acepta CSV o Excel (.xlsx/.xls) — requiere openpyxl para Excel.
    """
    path = file_meta.path
    if not path or not os.path.exists(path):
        raise RuntimeError("No se pudo acceder al archivo seleccionado.")
    ext = os.path.splitext(path)[1].lower()
    rows = []

    if ext == ".csv":
        with open(path, "r", encoding="utf-8-sig", newline="") as f:
            dr = csv.reader(f)
            headers = next(dr, None)
            if not headers:
                raise ValueError("El CSV no contiene encabezados.")
            i_code, i_name, i_desc, i_qty = _map_headers(headers)
            for r in dr:
                if not r or len(r) <= max(i_code, i_name):
                    continue
                code = normalize_string(r[i_code] if i_code >= 0 and i_code < len(r) else "")
                name = normalize_string(r[i_name] if i_name >= 0 and i_name < len(r) else "")
                desc = normalize_string(r[i_desc] if i_desc >= 0 and i_desc < len(r) else "")
                qty = to_int_safe(r[i_qty]) if (i_qty >= 0 and i_qty < len(r)) else 0
                if not (code and name):
                    continue
                rows.append({"code": code, "name": name, "description": desc, "qty": qty})

    elif ext in (".xlsx", ".xls"):
        try:
            from openpyxl import load_workbook
        except Exception:
            raise ImportError("Para archivos Excel instala: pip install openpyxl")
        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        header_cells = next(ws.iter_rows(min_row=1, max_row=1))
        headers = [("" if c.value is None else str(c.value)) for c in header_cells]
        i_code, i_name, i_desc, i_qty = _map_headers(headers)
        for row in ws.iter_rows(min_row=2):
            vals = [("" if c.value is None else str(c.value)) for c in row]
            if not vals or len(vals) <= max(i_code, i_name):
                continue
            code = normalize_string(vals[i_code] if i_code >= 0 and i_code < len(vals) else "")
            name = normalize_string(vals[i_name] if i_name >= 0 and i_name < len(vals) else "")
            desc = normalize_string(vals[i_desc] if i_desc >= 0 and i_desc < len(vals) else "")
            qty = to_int_safe(vals[i_qty]) if (i_qty >= 0 and i_qty < len(vals)) else 0
            if not (code and name):
                continue
            rows.append({"code": code, "name": name, "description": desc, "qty": qty})
    else:
        raise ValueError("Formato no soportado. Usa CSV o Excel (.xlsx/.xls).")

    if not rows:
        raise ValueError("No se encontraron filas válidas (requiere al menos Código y Nombre).")
    return rows


# -------------------------
# Acceso de datos (db)
# -------------------------
def existing_codes_set(db):
    codes = set()
    try:
        if hasattr(db, "list_products"):
            for p in db.list_products():
                c = normalize_string(str(p.get("code") or p.get("codigo") or ""))
                if c:
                    codes.add(c)
    except Exception:
        pass
    return codes


def fetch_products_for_warehouse(db, warehouse_id: int | None):
    """
    Devuelve (items, warn_msg)
    items: lista de productos (dicts) según el almacén o todos.
    """
    def _all_products():
        items = []
        try:
            if hasattr(db, "list_products"):
                items = db.list_products()
        except Exception:
            items = []
        return items

    if warehouse_id is None:
        return _all_products(), None

    if hasattr(db, "list_products_by_warehouse"):
        try:
            return db.list_products_by_warehouse(warehouse_id), None
        except Exception as ex:
            return _all_products(), f"No se pudo listar por almacén (usando todos): {ex}"

    return _all_products(), "Tu DB no expone list_products_by_warehouse; mostrando todos."


def build_stock_indexes(db):
    """
    Retorna:
    totals: dict[code] -> total en todos los almacenes
    per_wh: dict[warehouse_id] -> dict[code] -> qty en ese almacén
    wh_names: dict[warehouse_id] -> nombre almacén
    """
    totals: dict[str, int] = {}
    per_wh: dict[int, dict[str, int]] = {}
    wh_names: dict[int, str] = {}
    try:
        for w in db.list_warehouses():
            wid = int(w["id"])
            wh_names[wid] = w["name"]
            cmap: dict[str, int] = {}
            for p in db.list_products_by_warehouse(wid):
                code = str(p["code"])
                try:
                    q = int(p.get("qty") or 0)
                except Exception:
                    q = 0
                if q < 0:
                    q = 0
                cmap[code] = q
                totals[code] = totals.get(code, 0) + q
            per_wh[wid] = cmap
    except Exception:
        pass
    return totals, per_wh, wh_names


def get_stock_map(db, warehouse_id: int) -> dict:
    mp = {}
    try:
        for p in db.list_products_by_warehouse(warehouse_id):
            try:
                q = int(p.get("qty") or 0)
            except Exception:
                q = 0
            if q < 0:
                q = 0
            mp[str(p["code"])] = q
    except Exception:
        mp = {}
    return mp


# -------------------------
# Catálogo y búsqueda
# -------------------------
def search_collect_catalog(db, warehouse_id: int | None):
    """
    Construye el catálogo base (code, name, descr, total, wh_qty) dependiendo de si se filtra por almacén.
    """
    totals, per_wh, _ = build_stock_indexes(db)

    if warehouse_id is None:
        base = []
        try:
            base = db.list_products() or []
        except Exception:
            base = []
        seen = set()
        out = []
        for p in base:
            c = str(p.get("code") or "")
            if not c:
                continue
            seen.add(c)
            name = str(p.get("name") or p.get("nombre") or "")
            descr = str(p.get("description") or p.get("descripcion") or "")
            total_q = int(totals.get(c, 0))
            out.append({"code": c, "name": name, "descr": descr, "total": total_q, "wh_qty": None})
        for c, t in totals.items():
            if c not in seen:
                out.append({"code": c, "name": c, "descr": "", "total": int(t or 0), "wh_qty": None})
        return out
    else:
        try:
            items = db.list_products_by_warehouse(warehouse_id) or []
        except Exception:
            items = []
        cmap = per_wh.get(int(warehouse_id), {}) if per_wh else {}
        out = []
        for p in items:
            c = str(p.get("code") or "")
            if not c:
                continue
            name = str(p.get("name") or p.get("nombre") or "")
            descr = str(p.get("description") or p.get("descripcion") or "")
            total_q = int(totals.get(c, 0))
            wh_q = int(cmap.get(c, 0))
            out.append({"code": c, "name": name, "descr": descr, "total": total_q, "wh_qty": wh_q})
        return out


def search_filter_and_score(
    items: list[dict],
    query: str,
    include_descr: bool,
    in_stock_only: bool,
    low_only: bool,
    threshold: int,
    warehouse_id: int | None,
):
    """
    Filtra y ordena resultados devolviendo lista de items.
    """
    q = norm_text(query)
    thr = int(threshold or 5)

    def match_score(it):
        code = norm_text(it["code"])
        name = norm_text(it["name"])
        descr = norm_text(it.get("descr", ""))
        total_q = int(it.get("total") or 0)
        wh_q = it.get("wh_qty")
        qty = int(wh_q if warehouse_id is not None else total_q)

        if in_stock_only and qty <= 0:
            return None
        if low_only and not (0 < qty <= thr):
            return None

        if not q:
            return (1000 - min(qty, 999))

        score = 1000
        if code.startswith(q):
            score = min(score, 0)
        elif name.startswith(q):
            score = min(score, 1)
        elif q in code:
            score = min(score, 2)
        elif q in name:
            score = min(score, 3)
        elif include_descr and q in descr:
            score = min(score, 5)
        else:
            return None

        score += max(0, 20 - qty) // 4
        return score

    scored = []
    for it in items:
        sc = match_score(it)
        if sc is not None:
            scored.append((sc, it))
    scored.sort(key=lambda x: (x[0], x[1]["name"]))
    return [it for _, it in scored]


def now_timestamp_compact():
    # YYYYMMDD_HHMMSS
    from datetime import datetime
    return datetime.now().strftime("%Y%m%d_%H%M%S")
