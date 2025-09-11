import flet as ft
import database as db
import csv
import os
import threading


def main(page: ft.Page):
    db.init_db()
    db.ensure_color_column()
    db.ensure_products_table()

    # PROPIEDADES
    page.title = "Control de Almacen"
    page.padding = 0
    page.spacing = 0
    page.horizontal_alignment = "stretch"
    page.vertical_alignment = "stretch"
    page.theme_mode = ft.ThemeMode.LIGHT
    page.window.maximized = True
    last_import_rows: list[dict] = []

    warehouse_to_delete = {"id": None, "name": ""}
    ui_state = {
        "last_warehouse_id": None,     # último almacén mostrado en "Productos"
        "pending_file": None,          # archivo pendiente de importar
        "selected_wh_id": None,        # almacén elegido para importar
        "replace_stock": False,        # NUEVO: modo reemplazo de existencias
    }

    # ---- Entrada / Salida (estado) ----
    entry_state = {
        "warehouse_id": None,
        "lines": {},  # code -> {"name": str, "qty": int}
    }
    exit_state = {
        "warehouse_id": None,
        "lines": {},  # code -> {"name": str, "qty": int}
    }

    # Sobre-extracción (estado para diálogo)
    exit_over_state = {
        "warehouse_id": None,
        "items": [],  # [{"code","name","req","avail"}]
    }

    COLOR_CHOICES = [
        ("slate",  ["#0f172a", "#1f2937"]),
        ("indigo", ["#1e1b4b", "#312e81"]),
        ("emerald",["#064e3b", "#065f46"]),
        ("rose",   ["#7f1d1d", "#9f1239"]),
        ("amber",  ["#78350f", "#92400e"]),
        ("zinc",   ["#18181b", "#27272a"]),
    ]

    appbar_text_ref = ft.Ref[ft.Text]()

    # ====== Helper de SnackBars elegantes ======
    def notify(kind: str, message: str):
        """
        kind: 'success' | 'info' | 'warning' | 'error'
        """
        kind = (kind or "info").lower()
        colors = {
            "success": ft.Colors.GREEN_600,
            "info": ft.Colors.BLUE_600,
            "warning": ft.Colors.AMBER_700,
            "error": ft.Colors.RED_600,
        }
        icons = {
            "success": ft.Icons.CHECK_CIRCLE_ROUNDED,
            "info": ft.Icons.INFO_ROUNDED,
            "warning": ft.Icons.WARNING_AMBER_ROUNDED,
            "error": ft.Icons.ERROR_ROUNDED,
        }
        bg = colors.get(kind, ft.Colors.BLUE_600)
        ic = icons.get(kind, ft.Icons.INFO_ROUNDED)

        sb = ft.SnackBar(
            content=ft.Row(
                controls=[
                    ft.Icon(ic, color=ft.Colors.WHITE, size=22),
                    ft.Text(message, color=ft.Colors.WHITE, weight=ft.FontWeight.W_600),
                ],
                spacing=10,
                tight=True,
            ),
            bgcolor=bg,
            behavior=ft.SnackBarBehavior.FLOATING,
        )
        page.open(sb)

    # ===== Control único de diálogos (sin page.dialog) =====
    current_dialog = {"dlg": None}

    def open_dialog(dlg: ft.AlertDialog):
        if current_dialog["dlg"] is not None and getattr(current_dialog["dlg"], "open", False):
            current_dialog["dlg"].open = False
            page.update()
        current_dialog["dlg"] = dlg
        page.open(dlg)

    def close_dialog():
        if current_dialog["dlg"] is not None and getattr(current_dialog["dlg"], "open", False):
            current_dialog["dlg"].open = False
            page.update()
        current_dialog["dlg"] = None

    # ======= Paginación =======
    pagination_state = {"page": 0, "per_page": 100, "items": []}

    # ========================== Helpers de stock/productos ==========================
    def fetch_products_for_warehouse(warehouse_id: int | None):
        # Todos
        def _all_products():
            items = []
            try:
                if hasattr(db, "list_products"):
                    items = db.list_products()
            except Exception:
                items = []
            return items

        if warehouse_id is None:
            return _all_products(), None

        if hasattr(db, "list_products_by_warehouse"):
            try:
                return db.list_products_by_warehouse(warehouse_id), None
            except Exception as ex:
                return _all_products(), f"No se pudo listar por almacén (usando todos): {ex}"

        return _all_products(), "Tu DB no expone list_products_by_warehouse; mostrando todos."

    def build_stock_indexes():
        """
        Devuelve:
          totals: {code -> total en todos los almacenes}
          per_wh: {warehouse_id -> {code -> qty}}
          wh_names: {warehouse_id -> name}
        """
        totals: dict[str, int] = {}
        per_wh: dict[int, dict[str, int]] = {}
        wh_names: dict[int, str] = {}
        try:
            for w in db.list_warehouses():
                wid = int(w["id"])
                wh_names[wid] = w["name"]
                cmap: dict[str, int] = {}
                for p in db.list_products_by_warehouse(wid):
                    code = str(p["code"])
                    try:
                        q = int(p.get("qty") or 0)
                    except:
                        q = 0
                    if q < 0:
                        q = 0
                    cmap[code] = q
                    totals[code] = totals.get(code, 0) + q
                per_wh[wid] = cmap
        except Exception:
            pass
        return totals, per_wh, wh_names

    def open_product_detail(code: str, name: str):
        """Muestra un Alert con datos del producto y los almacenes donde se encuentra."""
        # Buscar descripción (si existe)
        descr = ""
        try:
            for p in (db.list_products() or []):
                if str(p.get("code")) == code:
                    descr = str(p.get("description") or p.get("descripcion") or "")
                    break
        except Exception:
            descr = ""

        totals, per_wh, wh_names = build_stock_indexes()
        total = int(totals.get(code, 0))
        rows = []
        for wid, cmap in per_wh.items():
            qty = int(cmap.get(code, 0))
            if qty > 0:
                rows.append(
                    ft.Row(
                        alignment=ft.MainAxisAlignment.SPACE_BETWEEN,
                        controls=[
                            ft.Text(wh_names.get(wid, "Almacén"), size=12),
                            ft.Text(str(qty), size=12, weight=ft.FontWeight.W_600),
                        ],
                    )
                )
        if not rows:
            rows = [ft.Text("No se encuentra en ningún almacén.", size=12, italic=True, color=ft.Colors.GREY_700)]

        dlg = ft.AlertDialog(
            modal=True,
            title=ft.Text(f"Producto: {code}"),
            shape=ft.RoundedRectangleBorder(radius=8),
            content=ft.Column(
                spacing=8,
                width=520,
                height=300,
                scroll=ft.ScrollMode.AUTO,
                controls=[
                    ft.Text(f"Nombre: {name}", size=12),
                    ft.Text(f"Descripción: {descr or '—'}", size=12),
                    ft.Divider(),
                    ft.Text(
                        f"Total existencias: {total}",
                        size=12,
                        weight=ft.FontWeight.W_600,
                        color=ft.Colors.RED_600 if total == 0 else None,
                    ),
                    ft.Text("Almacenes:", size=12, color=ft.Colors.GREY_700),
                    ft.Column(rows, spacing=6),
                ],
            ),
            actions=[
                ft.TextButton(
                    "Cerrar",
                    on_click=lambda e: (setattr(dlg, "open", False), page.update(), close_dialog()),
                )
            ],
            actions_alignment=ft.MainAxisAlignment.END,
        )
        open_dialog(dlg)

    def render_products_list(warehouse_id: int | None = None):
        # Índices de stock global y por almacén
        totals, per_wh, wh_names = build_stock_indexes()

        items, warn = fetch_products_for_warehouse(warehouse_id)
        if not items and last_import_rows:
            items = last_import_rows

        norm_items = []
        for it in items:
            code = str(it.get("code") or it.get("codigo") or it.get("Código") or "")
            name = str(it.get("name") or it.get("nombre") or it.get("Nombre") or "")
            # Total global
            total_q = int(totals.get(code, 0))
            # Qty del almacén actual (si aplica)
            wh_q = None
            if warehouse_id is not None:
                wh_q = int(per_wh.get(int(warehouse_id), {}).get(code, 0))
            norm_items.append({"code": code, "name": name, "total": total_q, "wh_qty": wh_q})
        # Si estamos en vista "todos los productos" y items venía desde list_products(),
        # puede que haya códigos que no aparezcan en 'items' pero sí en totales (por vínculos).
        # Añadimos faltantes para coherencia visual.
        if warehouse_id is None:
            for c, t in totals.items():
                if not any(x["code"] == c for x in norm_items):
                    # Intentar obtener nombre desde list_products()
                    pname = c
                    try:
                        for p in (db.list_products() or []):
                            if str(p.get("code")) == c:
                                pname = str(p.get("name") or p.get("nombre") or c)
                                break
                    except Exception:
                        pass
                    norm_items.append({"code": c, "name": pname, "total": int(t or 0), "wh_qty": None})

        norm_items.sort(key=lambda x: x["code"])

        pagination_state["items"] = norm_items
        pagination_state["page"] = 0

        wh_title = ""
        if warehouse_id is not None:
            ui_state["last_warehouse_id"] = warehouse_id
            try:
                ws = {w["id"]: w for w in db.list_warehouses()}
                if warehouse_id in ws:
                    wh_title = f" – {ws[warehouse_id]['name']}"
            except Exception:
                pass

        def build_table_page():
            per = pagination_state["per_page"]
            page_idx = pagination_state["page"]
            data = pagination_state["items"]
            total = len(data)
            start = page_idx * per
            end = min(start + per, total)

            rows = []
            for it in data[start:end]:
                code = it["code"]
                name = it["name"]
                total_q = int(it.get("total") or 0)
                wh_q = it.get("wh_qty")

                cells = [
                    ft.DataCell(ft.Container(content=ft.Text(code), padding=ft.padding.symmetric(6, 10))),
                    ft.DataCell(ft.Container(content=ft.Text(name), padding=ft.padding.symmetric(6, 10))),
                    ft.DataCell(
                        ft.Container(
                            content=ft.Text(
                                str(total_q),
                                color=ft.Colors.RED_600 if total_q == 0 else None
                            ),
                            padding=ft.padding.symmetric(6, 10),
                        )
                    ),
                ]
                if warehouse_id is not None:
                    wq = int(wh_q or 0)
                    cells.append(
                        ft.DataCell(
                            ft.Container(
                                content=ft.Text(
                                    str(wq),
                                    color=ft.Colors.RED_600 if wq == 0 else None
                                ),
                                padding=ft.padding.symmetric(6, 10),
                            )
                        )
                    )

                def on_row_click(e, c=code, n=name):
                    open_product_detail(c, n)
                    # opcional: quitar el highlight de selección
                    try:
                        e.control.selected = False
                        page.update()
                    except:
                        pass

                rows.append(
                    ft.DataRow(
                        cells=cells,
                        on_select_changed=on_row_click,   # click en toda la fila
                    )
                )

            columns = [
                ft.DataColumn(ft.Text("Código")),
                ft.DataColumn(ft.Text("Nombre")),
                ft.DataColumn(ft.Text("Total")),
            ]
            if warehouse_id is not None:
                columns.append(ft.DataColumn(ft.Text("Existencia")))

            table = ft.DataTable(
                columns=columns,
                rows=rows,
                column_spacing=20,
                heading_row_height=40,
                data_row_max_height=56,
                show_checkbox_column=False,   # sin checkbox, pero la fila es clickeable
                # ⚠️ quitamos data_row_color para evitar el AttributeError
            )

            total_pages = max(1, (total + per - 1) // per)
            page_label = ft.Text(f"Página {page_idx + 1} de {total_pages} • {total} producto(s)")

            def go_prev(e):
                if pagination_state["page"] > 0:
                    pagination_state["page"] -= 1
                    render_controls()

            def go_next(e):
                if (pagination_state["page"] + 1) * per < total:
                    pagination_state["page"] += 1
                    render_controls()

            pager = ft.Row(
                alignment=ft.MainAxisAlignment.END,
                controls=[
                    ft.FilledButton("Anterior", on_click=go_prev, disabled=(page_idx == 0)),
                    ft.FilledButton("Siguiente", on_click=go_next, disabled=(end >= total)),
                ],
            )

            return table, page_label, pager



        def render_controls():
            table, page_label, pager = build_table_page()
            header_text = "Productos" + wh_title

            # ----- BOTONES SUPERIORES (derecha) -----
            top_actions = []
            if warehouse_id is not None:
                top_actions = [
                    ft.FilledTonalButton("← Almacenes", icon=ft.Icons.ARROW_BACK,
                                         on_click=lambda e: render_warehouses()),
                    ft.FilledButton("Entrada (escáner)", icon=ft.Icons.QR_CODE,
                                    on_click=lambda e, wid=warehouse_id: open_entry_for(wid)),
                    ft.FilledButton("Salida", icon=ft.Icons.EXIT_TO_APP,
                                    on_click=lambda e, wid=warehouse_id: open_exit_for(wid)),
                ]

            header_row = ft.Row(
                alignment=ft.MainAxisAlignment.SPACE_BETWEEN,
                vertical_alignment=ft.CrossAxisAlignment.CENTER,
                controls=[
                    ft.Text(header_text, size=20, weight=ft.FontWeight.BOLD),
                    ft.Row(controls=top_actions, spacing=8),
                ],
            )

            extra = []
            if warn:
                extra.append(ft.Text(warn, size=11, color=ft.Colors.ORANGE_700))

            content_column.controls[:] = [
                ft.Container(padding=ft.padding.only(8, 0, 8, 8),
                             content=ft.Column(controls=[header_row] + extra, spacing=6)),
                ft.Container(padding=ft.padding.only(8, 0, 8, 4), content=page_label),
                ft.Container(expand=True, padding=ft.padding.all(8), content=ft.ListView(expand=True, controls=[table])),
                ft.Container(padding=ft.padding.all(8), content=pager),
            ]
            page.update()

        render_controls()

    # Helper: refrescar la vista actual de productos (si está filtrada por almacén)
    def refresh_current_products_view():
        wid = ui_state.get("last_warehouse_id")
        if wid is not None:
            render_products_list(wid)

    # -------- Render de almacenes 300x300 --------
    def render_warehouses():
        warehouses = db.list_warehouses()
        if not warehouses:
            content_column.controls[:] = [
                ft.Container(
                    padding=20,
                    content=ft.Text("Aún no hay almacenes. Ve a Almacén > Crear un almacén.", size=16)
                )
            ]
            page.update()
            return

        cards = []
        for w in warehouses:
            def _open_wh_products(e, _w=w):
                render_products_list(_w["id"])

            card_body = ft.Container(
                width=300,
                height=300,
                gradient=gradient_for(w.get("color_key") or "slate"),
                border_radius=8,
                padding=16,
                ink=True,
                on_click=_open_wh_products,
                shadow=ft.BoxShadow(spread_radius=1, blur_radius=18, color=ft.Colors.BLACK26, offset=ft.Offset(0, 6)),
                content=ft.Column(
                    alignment=ft.MainAxisAlignment.CENTER,
                    horizontal_alignment=ft.CrossAxisAlignment.CENTER,
                    controls=[
                        ft.Icon(ft.Icons.WAREHOUSE, size=68, color=ft.Colors.WHITE),
                        ft.Text(w["name"], size=18, weight=ft.FontWeight.BOLD, text_align=ft.TextAlign.CENTER, color=ft.Colors.WHITE),
                        ft.Text((w.get("description") or "—"), size=12, color=ft.Colors.WHITE70, text_align=ft.TextAlign.CENTER),
                    ],
                ),
            )

            delete_btn = ft.IconButton(
                icon=ft.Icons.DELETE_FOREVER_ROUNDED,
                icon_size=20, width=25, height=25, tooltip="Eliminar almacén",
                icon_color=ft.Colors.RED_600,
                style=ft.ButtonStyle(shape=ft.RoundedRectangleBorder(radius=5), padding=2, bgcolor=ft.Colors.WHITE),
                on_click=lambda e, _w=w: confirm_delete(_w),
            )
            card = ft.Stack(width=300, height=300, controls=[card_body, ft.Container(right=8, top=8, content=delete_btn, border_radius=5)])
            cards.append(card)

        grid = ft.GridView(expand=True, runs_count=3, max_extent=320, child_aspect_ratio=1, spacing=16, run_spacing=16, controls=cards)

        content_column.controls[:] = [
            ft.Container(padding=ft.padding.only(8, 0, 8, 8), content=ft.Text("Almacenes", size=20, weight=ft.FontWeight.BOLD)),
            ft.Container(expand=True, padding=ft.padding.all(8), content=grid),
        ]
        page.update()

    # Ir directo a crear almacén desde el alerta de “no hay almacenes”
    def goto_create(e):
        close_dialog()
        open_create_dialog()

    # ========== IMPORTACIÓN con selección de ALMACÉN ==========
    def render_import_products():
        warehouses = db.list_warehouses()
        if not warehouses:
            open_dialog(alert_no_wh); return

        hint = ft.Column(
            spacing=10,
            horizontal_alignment=ft.CrossAxisAlignment.CENTER,
            controls=[
                ft.Icon(ft.Icons.FILE_UPLOAD, size=60, color=ft.Colors.GREY_700),
                ft.Text("Arrastra tu archivo CSV o Excel aquí", size=16, weight=ft.FontWeight.BOLD),
                ft.Text("o haz clic para seleccionar un archivo", size=12, color=ft.Colors.GREY_600),
                ft.FilledButton("Cargar archivo", icon=ft.Icons.UPLOAD_FILE,
                    on_click=lambda e: file_picker.pick_files(allow_multiple=False, allowed_extensions=["csv","CSV","xlsx","XLSX","xls","XLS"])),
                ft.Text("Campos requeridos: Código, Nombre; Descripción y Existencias (opcional)", size=11, color=ft.Colors.GREY_700),
            ],
        )

        drop_zone = ft.Container(
            width=520, height=280, bgcolor=ft.Colors.GREY_100,
            border=ft.border.all(2, ft.Colors.GREY_300), border_radius=12, ink=True,
            on_click=lambda e: file_picker.pick_files(allow_multiple=False, allowed_extensions=["csv","xlsx","xls"]),
            content=hint, alignment=ft.alignment.center,
        )

        content_column.controls[:] = [ft.Container(expand=True, alignment=ft.alignment.center, content=drop_zone)]
        page.update()

    def open_create_dialog():
        name_tf.value = ""; descr_tf.value = ""
        dlg_create.open = True
        page.open(dlg_create)

    def save_warehouse():
        name = (name_tf.value or "").strip()
        desc = (descr_tf.value or "").strip()
        if not name:
            notify("warning", "El nombre es obligatorio."); return
        try:
            db.add_warehouse(name=name, description=desc, color_key=color_dd.value)
            notify("success", f"Almacén '{name}' creado.")
            dlg_create.open = False; page.update(); close_dialog()
            render_warehouses(); page.update()
        except Exception as ex:
            notify("error", f"Error: {ex}")

    # ==== Handlers de menú ====
    def handle_submenu_open(e: ft.ControlEvent): pass
    def handle_submenu_close(e: ft.ControlEvent): pass
    def handle_submenu_hover(e: ft.ControlEvent): pass

    def handle_menu_item_click(e: ft.ControlEvent):
        cmd = getattr(e.control, "data", None)
        label = e.control.content.value if hasattr(e.control, "content") else ""

        if cmd == "warehouses_view" or label == "Ver almacenes":
            render_warehouses()
        elif cmd == "warehouse_new" or label == "Crear un almacen":
            open_create_dialog()
        elif label == "Entrada de productos":
            open_entry_dialog()
        elif cmd == "products_import" or label == "Agregar Lista de Productos":
            render_import_products()
        elif cmd == "products_view" or label == "Ver productos":
            ui_state["last_warehouse_id"] = None
            render_products_list(None)
        else:
            notify("info", f"{label} was clicked!")

        if appbar_text_ref.current:
            appbar_text_ref.current.value = label
            page.update()

    # ======= Helpers de importación / duplicados =========
    def _normalize(s: str) -> str: return (s or "").strip()

    def _to_int_safe(v) -> int:
        try:
            s = str(v).strip()
            if s == "" or s.lower() in ("none", "nan", "null"):
                return 0
            return int(float(s.replace(",", "")))
        except:
            return 0

    def _existing_codes_set():
        codes = set()
        try:
            if hasattr(db, "list_products"):
                for p in db.list_products():
                    c = _normalize(str(p.get("code") or p.get("codigo") or ""))
                    if c: codes.add(c)
        except Exception: pass
        return codes

    def _map_headers(headers: list[str]):
        def _norm(h: str):
            h = (h or "").strip().lower()
            for a,b in [("á","a"),("é","e"),("í","i"),("ó","o"),("ú","u")]: h = h.replace(a,b)
            return h
        norm = [_norm(h) for h in headers]
        def find(*c):
            for k in c:
                if k in norm: return norm.index(k)
            return -1
        i_code = find("codigo","código","code","sku","clave")
        i_name = find("nombre","name","producto")
        i_desc = find("descripcion","descripción","description","detalle")
        i_qty  = find("existencia","existencias","qty","cantidad","stock","inventario","exist")
        if i_code < 0 or i_name < 0:
            raise ValueError("Encabezados requeridos: Código y Nombre (Descripción y Existencias opcionales).")
        return i_code, i_name, i_desc, i_qty

    def parse_products_from_file(file_meta) -> list[dict]:
        path = file_meta.path
        if not path or not os.path.exists(path):
            raise RuntimeError("No se pudo acceder al archivo seleccionado.")
        ext = os.path.splitext(path)[1].lower()
        rows = []

        if ext == ".csv":
            with open(path, "r", encoding="utf-8-sig", newline="") as f:
                dr = csv.reader(f); headers = next(dr, None)
                if not headers: raise ValueError("El CSV no contiene encabezados.")
                i_code, i_name, i_desc, i_qty = _map_headers(headers)
                for r in dr:
                    if not r or len(r) <= max(i_code, i_name): continue
                    code = _normalize(r[i_code] if i_code >=0 and i_code < len(r) else "")
                    name = _normalize(r[i_name] if i_name >=0 and i_name < len(r) else "")
                    desc = _normalize(r[i_desc] if i_desc >=0 and i_desc < len(r) else "")
                    qty  = _to_int_safe(r[i_qty]) if (i_qty >= 0 and i_qty < len(r)) else 0
                    if not (code and name): continue
                    rows.append({"code": code, "name": name, "description": desc, "qty": qty})
        elif ext in (".xlsx", ".xls"):
            try:
                from openpyxl import load_workbook
            except Exception:
                raise ImportError("Para archivos Excel instala: pip install openpyxl")
            wb = load_workbook(path, read_only=True, data_only=True)
            ws = wb.active
            header_cells = next(ws.iter_rows(min_row=1, max_row=1))
            headers = [("" if c.value is None else str(c.value)) for c in header_cells]
            i_code, i_name, i_desc, i_qty = _map_headers(headers)
            for row in ws.iter_rows(min_row=2):
                vals = [("" if c.value is None else str(c.value)) for c in row]
                if not vals or len(vals) <= max(i_code, i_name): continue
                code = _normalize(vals[i_code] if i_code >=0 and i_code < len(vals) else "")
                name = _normalize(vals[i_name] if i_name >=0 and i_name < len(vals) else "")
                desc = _normalize(vals[i_desc] if i_desc >=0 and i_desc < len(vals) else "")
                qty  = _to_int_safe(vals[i_qty]) if (i_qty >= 0 and i_qty < len(vals)) else 0
                if not (code and name): continue
                rows.append({"code": code, "name": name, "description": desc, "qty": qty})
        else:
            raise ValueError("Formato no soportado. Usa CSV o Excel (.xlsx/.xls).")

        if not rows:
            raise ValueError("No se encontraron filas válidas (requiere al menos Código y Nombre).")
        return rows

    # ======= Importación con selección de almacén =====
    def import_rows_with_progress(rows: list[dict], warehouse_id: int, replace_mode: bool = False):
        existing = _existing_codes_set()
        preprocessed = [{
            "code": _normalize(r["code"]),
            "name": _normalize(r["name"]),
            "description": _normalize(r.get("description","")),
            "qty": int(r.get("qty") or 0)
        } for r in rows]

        total = len(preprocessed)
        prog = ft.ProgressBar(value=0, width=400)
        lbl  = ft.Text(f"0 / {total} productos", size=12)

        dlg_prog = ft.AlertDialog(
            modal=True,
            title=ft.Text("Cargando productos..."),
            content=ft.Column([prog, lbl], spacing=10, width=420, height=60),
            shape=ft.RoundedRectangleBorder(radius=5),
            shadow_color=ft.Colors.BLACK38,
            actions=[],
        )
        open_dialog(dlg_prog)

        # Cache de existencias actuales del almacén (para fallback de reemplazo)
        current_stock = {}
        try:
            for p in db.list_products_by_warehouse(warehouse_id):
                current_stock[str(p["code"])] = int(p.get("qty") or 0)
        except Exception:
            current_stock = {}

        def _link_to_warehouse(code: str, wid: int):
            try:
                if hasattr(db, "is_product_linked") and db.is_product_linked(code, wid):
                    return
            except Exception:
                pass
            if hasattr(db, "link_product_to_warehouse"):
                db.link_product_to_warehouse(code, wid)
            else:
                db.upsert_product(code=code, name=None, description=None, warehouse_id=wid)

        def _set_or_add_stock(code: str, wid: int, qty: int):
            if qty <= 0:
                return
            if replace_mode:
                # 1) Si existe set_stock(), úsalo
                if hasattr(db, "set_stock"):
                    try:
                        db.set_stock(code, wid, qty, note="Importación (reemplazo)")
                        # actualizar cache
                        current_stock[code] = qty
                        return
                    except Exception:
                        pass
                # 2) Fallback: calcula delta y aplica increment/decrement
                prev = current_stock.get(code, 0)
                delta = qty - prev
                if delta > 0:
                    db.increment_stock(code, wid, delta, note="Importación (ajuste a reemplazo +)")
                elif delta < 0:
                    db.decrement_stock(code, wid, -delta, note="Importación (ajuste a reemplazo -)")
                current_stock[code] = qty
            else:
                # Sumar
                db.increment_stock(code, wid, qty, note="Importación (suma)")

        def worker():
            ok = link_ok = err = 0
            for i, r in enumerate(preprocessed, start=1):
                try:
                    code_key = r["code"]
                    # Alta/actualización del catálogo base
                    if code_key not in existing:
                        db.upsert_product(
                            code=r["code"], name=r["name"], description=r.get("description",""), warehouse_id=None
                        )
                        existing.add(code_key)
                        ok += 1

                    # Vincular al almacén
                    _link_to_warehouse(code_key, warehouse_id); link_ok += 1

                    # Aplicar existencias si vienen en archivo
                    _set_or_add_stock(code_key, warehouse_id, int(r.get("qty") or 0))

                except Exception:
                    err += 1

                prog.value = i / total
                lbl.value = f"{i} / {total} productos"
                page.update()

            dlg_prog.open = False; page.update(); close_dialog()
            modo = "reemplazo" if replace_mode else "suma"
            msg = f"Importación ({modo}) completada: {ok} nuevos, {link_ok} asociados"
            if err: msg += f", {err} con error"
            notify("success", msg)
            render_products_list(warehouse_id)

        threading.Thread(target=worker, daemon=True).start()

    def process_selected_file_with_warehouse(file_meta, warehouse_id: int, replace_mode: bool = False):
        try:
            rows = parse_products_from_file(file_meta)
        except ImportError as ie:
            notify("error", str(ie)); return
        except Exception as ex:
            notify("error", f"Error al leer archivo: {ex}"); return
        import_rows_with_progress(rows, warehouse_id, replace_mode)

    # ---------- FilePicker ----------
    def on_file_picked(e: ft.FilePickerResultEvent):
        if not e.files: return
        f = e.files[0]
        ui_state["pending_file"] = f
        refresh_pick_wh_dialog_and_open()

    # ====== Diálogo: Seleccionar almacén para importar ======
    def refresh_pick_wh_dialog_and_open():
        warehouses = db.list_warehouses()
        if not warehouses:
            open_dialog(alert_no_wh); return
        wh_options = [ft.dropdown.Option(str(w["id"]), text=w["name"]) for w in warehouses]
        wh_dd.options = wh_options
        wh_dd.value = wh_options[0].key if wh_options else None
        ui_state["selected_wh_id"] = int(wh_dd.value) if wh_dd.value else None
        replace_stock_cb.value = bool(ui_state.get("replace_stock", False))
        open_dialog(dlg_pick_wh)

    def on_pick_wh_changed(e):
        val = wh_dd.value
        ui_state["selected_wh_id"] = int(val) if val is not None else None

    def on_pick_wh_confirm(e):
        if ui_state["pending_file"] is None or ui_state["selected_wh_id"] is None:
            notify("warning", "Selecciona un almacén válido."); return
        dlg_pick_wh.open = False; page.update(); close_dialog()
        process_selected_file_with_warehouse(
            ui_state["pending_file"],
            ui_state["selected_wh_id"],
            replace_mode=bool(ui_state.get("replace_stock", False)),
        )
        ui_state["pending_file"] = None

    def on_pick_wh_cancel(e):
        dlg_pick_wh.open = False; page.update(); close_dialog()
        ui_state["pending_file"] = None; ui_state["selected_wh_id"] = None

    # -------- ENTRADA DE PRODUCTOS (alert) --------
    entry_wh_dd = ft.Dropdown(label="Almacén", width=360)
    entry_code_tf = ft.TextField(
        label="Código / Escáner",
        autofocus=True,
        width=360,
        on_submit=lambda e: entry_add_code(e.control.value),
        keyboard_type=ft.KeyboardType.NUMBER
    )
    entry_lines_col = ft.Column(spacing=8, width=520, tight=True, scroll=ft.ScrollMode.ADAPTIVE)

    # ====== Helpers de foco ======
    def focus_entry_field():
        try:
            entry_code_tf.focus()
        except Exception:
            pass
        page.update()

    def focus_exit_field():
        try:
            exit_code_tf.focus()
        except Exception:
            pass
        page.update()

    def _entry_refresh_warehouse_options():
        ws = db.list_warehouses()
        entry_wh_dd.options = [ft.dropdown.Option(str(w["id"]), text=w["name"]) for w in ws]
        if ws:
            entry_wh_dd.value = str(ws[0]["id"])
            entry_state["warehouse_id"] = ws[0]["id"]
        else:
            entry_wh_dd.value = None
            entry_state["warehouse_id"] = None

    def entry_on_wh_change(e):
        entry_state["warehouse_id"] = int(entry_wh_dd.value) if entry_wh_dd.value else None
        entry_state["lines"].clear()
        entry_render_lines()
        focus_entry_field()

    entry_wh_dd.on_change = entry_on_wh_change

    # --- NUEVO: confirmar cantidad con Enter (entrada) ---
    def entry_qty_submit(code: str, val: str):
        try:
            n = int(val)
            n = 1 if n <= 0 else n
        except:
            n = 1
        if code in entry_state["lines"]:
            entry_state["lines"][code]["qty"] = n
        # Regresar foco SOLO cuando confirma con Enter
        focus_entry_field()

    def entry_render_lines():
        rows = []
        for code, data in entry_state["lines"].items():
            qty_tf = ft.TextField(
                value=str(data["qty"]),
                width=70,
                text_align=ft.TextAlign.RIGHT,
                on_change=lambda e, _c=code: entry_update_qty(_c, e.control.value),
                on_submit=lambda e, _c=code: entry_qty_submit(_c, e.control.value),
                keyboard_type=ft.KeyboardType.NUMBER,
            )
            rows.append(
                ft.Row(
                    alignment=ft.MainAxisAlignment.SPACE_BETWEEN,
                    controls=[ft.Text(f"{code} – {data['name']}", size=13), ft.Row(controls=[ft.Text("Cant."), qty_tf])],
                )
            )
        entry_lines_col.controls = rows
        page.update()

    def entry_update_qty(code: str, val: str):
        try:
            n = int(val);  n = 1 if n <= 0 else n
        except: n = 1
        if code in entry_state["lines"]:
            entry_state["lines"][code]["qty"] = n
        # NO mover el foco aquí

    def entry_add_code(raw_code: str):
        code = (raw_code or "").strip()
        entry_code_tf.value = ""
        page.update()
        if not code:
            focus_entry_field(); return
        wid = entry_state["warehouse_id"]
        if not wid:
            notify("warning", "Selecciona un almacén."); focus_entry_field(); return
        try:
            prods = db.list_products_by_warehouse(wid)
            mp = {p["code"]: p for p in prods}
            if code not in mp:
                notify("warning", "El código no pertenece a este almacén."); focus_entry_field(); return
            name = mp[code]["name"]
        except Exception as ex:
            notify("error", f"No se pudo validar producto: {ex}"); focus_entry_field(); return

        entry_state["lines"].setdefault(code, {"name": name, "qty": 0})
        entry_state["lines"][code]["qty"] += 1
        entry_render_lines()
        focus_entry_field()

    def entry_confirm(e):
        wid = entry_state["warehouse_id"]
        if not wid or not entry_state["lines"]:
            notify("warning", "No hay productos que registrar."); focus_entry_field(); return
        total_items = 0
        for code, data in entry_state["lines"].items():
            qty = int(data["qty"] or 0)
            if qty > 0:
                db.increment_stock(code, wid, qty, note="Entrada manual")
                total_items += qty

        dlg_entry.open = False
        page.update()
        close_dialog()

        if ui_state.get("last_warehouse_id") == wid:
            render_products_list(wid)
            page.update()

        notify("success", f"Entrada registrada: {len(entry_state['lines'])} productos, {total_items} unidades.")
        entry_state["lines"].clear()

    def entry_clear(e):
        entry_state["lines"].clear()
        entry_render_lines()
        focus_entry_field()

    def open_entry_dialog():
        if not db.list_warehouses():
            open_dialog(alert_no_wh); return
        _entry_refresh_warehouse_options()
        entry_state["lines"].clear()
        entry_render_lines()
        page.open(dlg_entry)
        focus_entry_field()

    # ===== Helpers: abrir entrada/salida con almacén preseleccionado =====
    def open_entry_for(warehouse_id: int):
        if not db.list_warehouses():
            open_dialog(alert_no_wh); return
        _entry_refresh_warehouse_options()
        opts = {opt.key for opt in (entry_wh_dd.options or [])}
        if str(warehouse_id) in opts:
            entry_wh_dd.value = str(warehouse_id)
            entry_state["warehouse_id"] = warehouse_id
        else:
            entry_state["warehouse_id"] = int(entry_wh_dd.value) if entry_wh_dd.value else None
        entry_state["lines"].clear()
        entry_render_lines()
        page.open(dlg_entry)
        focus_entry_field()

    # -------- SALIDA DE PRODUCTOS (alert) --------
    def exit_on_wh_change(e):
        exit_state["warehouse_id"] = int(exit_wh_dd.value) if exit_wh_dd.value else None
        exit_render_lines()
        # NO mover el foco aquí para no interrumpir edición

    # --- NUEVO: confirmar cantidad con Enter (salida) ---
    def exit_qty_submit(code: str, val: str):
        try:
            n = int(val)
            n = 1 if n <= 0 else n
        except:
            n = 1
        if code in exit_state["lines"]:
            exit_state["lines"][code]["qty"] = n
        # Regresar foco SOLO cuando confirma con Enter
        focus_exit_field()

    def exit_render_lines():
        rows = []
        for code, data in exit_state["lines"].items():
            qty_tf = ft.TextField(
                value=str(data["qty"]),
                width=70,
                text_align=ft.TextAlign.RIGHT,
                on_change=lambda e, _c=code: exit_update_qty(_c, e.control.value),
                on_submit=lambda e, _c=code: exit_qty_submit(_c, e.control.value),
                keyboard_type=ft.KeyboardType.NUMBER,
            )
            rows.append(
                ft.Row(
                    alignment=ft.MainAxisAlignment.SPACE_BETWEEN,
                    controls=[ft.Text(f"{code} – {data['name']}", size=13), ft.Row(controls=[ft.Text("Cant."), qty_tf])],
                )
            )
        exit_lines_col.controls[:] = rows
        page.update()

    def exit_update_qty(code: str, val: str):
        try:
            n = int(val)
            n = 1 if n <= 0 else n
        except:
            n = 1
        if code in exit_state["lines"]:
            exit_state["lines"][code]["qty"] = n
        # IMPORTANTE: NO re-render ni foco aquí para no romper edición

    # ====== Helpers de stock ======
    def get_stock_map(warehouse_id: int) -> dict:
        """Devuelve {code: qty>=0} del almacén."""
        mp = {}
        try:
            for p in db.list_products_by_warehouse(warehouse_id):
                try:
                    q = int(p.get("qty") or 0)
                except:
                    q = 0
                if q < 0:
                    q = 0
                mp[str(p["code"])] = q
        except Exception:
            mp = {}
        return mp

    # ====== Diálogo de sobre-extracción ======
    exit_over_list_col = ft.Column(spacing=8, tight=True, width=520, height=220, scroll=ft.ScrollMode.AUTO)

    def apply_exit_caps_and_perform(e):
        """Ajusta cantidades a máximos y realiza la salida."""
        wid = exit_over_state.get("warehouse_id")
        for it in exit_over_state.get("items", []):
            code = it["code"]
            avail = int(it["avail"] or 0)
            if code in exit_state["lines"]:
                exit_state["lines"][code]["qty"] = avail  # cap
        dlg_exit_over.open = False
        page.update()
        close_dialog()
        perform_exit(wid)

    dlg_exit_over = ft.AlertDialog(
        modal=True,
        title=ft.Text("Cantidad solicitada supera la existencia"),
        content=ft.Column(
            controls=[
                ft.Text("Se encontraron productos con cantidad solicitada mayor que la existencia. ¿Deseas extraer la cantidad máxima disponible?", size=12),
                ft.Divider(),
                exit_over_list_col,
            ],
            spacing=8,
            width=560,
            height=320,
            scroll=ft.ScrollMode.AUTO,
        ),
        actions=[
            ft.TextButton("Cancelar", on_click=lambda e: (setattr(dlg_exit_over, "open", False), page.update(), close_dialog())),
            ft.FilledButton("Usar máximos y continuar", on_click=apply_exit_caps_and_perform),
        ],
        actions_alignment=ft.MainAxisAlignment.END,
        shape=ft.RoundedRectangleBorder(radius=5),
    )

    def show_exit_over_dialog(over_items: list[dict], wid: int):
        """Rellena y muestra el diálogo de sobre-extracción."""
        exit_over_state["items"] = over_items
        exit_over_state["warehouse_id"] = wid
        # Construir lista
        rows = []
        for it in over_items:
            rows.append(
                ft.Row(
                    alignment=ft.MainAxisAlignment.SPACE_BETWEEN,
                    controls=[
                        ft.Text(f"{it['code']} – {it['name']}", size=12),
                        ft.Text(f"Solicitado: {it['req']} • Máx: {it['avail']}", size=12, weight=ft.FontWeight.W_600),
                    ]
                )
            )
        exit_over_list_col.controls = rows
        open_dialog(dlg_exit_over)

    # ====== Flujo de salida ======
    def perform_exit(wid: int | None):
        """Aplica la salida (ya validada/capeada) y da feedback."""
        if not wid:
            notify("warning", "Selecciona un almacén."); focus_exit_field(); return
        if not exit_state["lines"]:
            notify("warning", "No hay productos que registrar."); focus_exit_field(); return

        total_items = 0
        for code, data in exit_state["lines"].items():
            qty = int(data["qty"] or 0)
            if qty > 0:
                db.decrement_stock(code, wid, qty, note="Salida manual")
                total_items += qty

        # Cierra el diálogo principal si aún estuviera abierto
        try:
            dlg_exit.open = False
        except:
            pass
        page.update()
        close_dialog()

        # Refresca la vista si corresponde
        if ui_state.get("last_warehouse_id") == wid:
            render_products_list(wid)
            page.update()

        notify("success", f"Salida registrada: {len(exit_state['lines'])} productos, {total_items} unidades.")
        exit_state["lines"].clear()

    def exit_add_code(raw_code: str):
        code = (raw_code or "").strip()
        exit_code_tf.value = ""
        page.update()

        if not code:
            focus_exit_field(); return

        wid = exit_state["warehouse_id"]
        if not wid:
            notify("warning", "Selecciona un almacén.")
            focus_exit_field(); return

        try:
            prods = db.list_products_by_warehouse(wid)
            mp = {p["code"]: p for p in prods}
            if code not in mp:
                notify("warning", "El código no pertenece a este almacén.")
                focus_exit_field(); return
            name = mp[code]["name"]
        except Exception as ex:
            notify("error", f"No se pudo validar producto: {ex}")
            focus_exit_field(); return

        if code in exit_state["lines"]:
            exit_state["lines"][code]["qty"] += 1
        else:
            exit_state["lines"][code] = {"name": name, "qty": 1}

        exit_render_lines()
        focus_exit_field()

    def exit_confirm(e):
        wid = exit_state["warehouse_id"]
        if not wid or not exit_state["lines"]:
            notify("warning", "No hay productos que registrar."); focus_exit_field(); return

        # Validar contra existencias actuales
        stock = get_stock_map(wid)
        over = []
        for code, data in exit_state["lines"].items():
            req = int(data.get("qty") or 0)
            avail = int(stock.get(code, 0))
            if req > avail:
                over.append({"code": code, "name": data.get("name",""), "req": req, "avail": avail})

        if over:
            # SnackBar y diálogo de confirmación de máximos
            notify("error", "No se puede extraer más productos que los que están en existencia.")
            show_exit_over_dialog(over, wid)
            return

        # Si no hay sobre-extracción, procesar normalmente
        perform_exit(wid)

    def exit_clear(e):
        exit_state["lines"].clear()
        exit_render_lines()
        focus_exit_field()

    def _exit_refresh_warehouse_options():
        ws = db.list_warehouses()
        exit_wh_dd.options = [ft.dropdown.Option(str(w["id"]), text=w["name"]) for w in ws]
        if ws:
            exit_wh_dd.value = str(ws[0]["id"])
            exit_state["warehouse_id"] = ws[0]["id"]
        else:
            exit_wh_dd.value = None
            exit_state["warehouse_id"] = None

    def open_exit_dialog(e):
        if not db.list_warehouses():
            open_dialog(alert_no_wh); return
        _exit_refresh_warehouse_options()
        exit_state["lines"].clear()
        exit_render_lines()
        page.open(dlg_exit)
        focus_exit_field()

    def open_exit_for(warehouse_id: int):
        if not db.list_warehouses():
            open_dialog(alert_no_wh); return
        _exit_refresh_warehouse_options()
        opts = {opt.key for opt in (exit_wh_dd.options or [])}
        if str(warehouse_id) in opts:
            exit_wh_dd.value = str(warehouse_id)
            exit_state["warehouse_id"] = warehouse_id
        else:
            exit_state["warehouse_id"] = int(exit_wh_dd.value) if exit_wh_dd.value else None
        exit_state["lines"].clear()
        exit_render_lines()
        page.open(dlg_exit)
        focus_exit_field()

    # ---------- UI base / common ----------
    file_picker = ft.FilePicker(on_result=on_file_picked)
    page.overlay.append(file_picker)

    exit_wh_dd = ft.Dropdown(label="Almacén", width=360)
    exit_code_tf = ft.TextField(
        label="Código / Escáner",
        autofocus=True,
        width=360,
        on_submit=lambda e: exit_add_code(e.control.value),
        keyboard_type=ft.KeyboardType.NUMBER,
        on_change=lambda e: exit_on_wh_change(e.control.value),
    )
    exit_lines_col = ft.Column(spacing=8, width=520, tight=True, scroll=ft.ScrollMode.ADAPTIVE)

    dlg_exit = ft.AlertDialog(
        modal=True,
        title=ft.Text("Salida de productos"),
        content=ft.Column(
            width=560, spacing=12,
            controls=[
                exit_wh_dd,
                exit_code_tf,
                ft.Divider(),
                ft.Text("Productos en esta salida:", size=12, color=ft.Colors.GREY_700),
                exit_lines_col,
            ], height=400, scroll=ft.ScrollMode.AUTO
        ), shape=ft.RoundedRectangleBorder(radius=5),
        actions=[
            ft.TextButton("Vaciar", on_click=exit_clear),
            ft.TextButton("Cancelar", on_click=lambda e: (setattr(dlg_exit, "open", False), page.update(), close_dialog())),
            ft.FilledButton("Confirmar", on_click=exit_confirm),
        ],
        actions_alignment=ft.MainAxisAlignment.END,
    )

    content_column = ft.Column(expand=True, scroll=None, horizontal_alignment=ft.CrossAxisAlignment.STRETCH)
    color_dd = ft.Dropdown(label="Color de tarjeta", width=400, value="slate",
                           options=[ft.dropdown.Option(k, text=k.capitalize()) for k,_ in COLOR_CHOICES])
    content_area = ft.Container(top=56, left=0, right=0, bottom=0, padding=ft.padding.only(20, 20, 20, 20), content=content_column)
    name_tf = ft.TextField(label="Nombre del almacén", autofocus=True, width=400)
    descr_tf = ft.TextField(label="Descripción (opcional)", width=400, height=100)

    dlg_delete = ft.AlertDialog(
        modal=True,
        title=ft.Text("Eliminar almacén"),
        content=ft.Text("Esta acción eliminará el almacén y sus datos relacionados (vínculos/stock). ¿Deseas continuar?"),
        actions=[
            ft.TextButton("Cancelar", on_click=lambda e: (setattr(dlg_delete, "open", False), page.update(), close_dialog())),
            ft.FilledButton("Eliminar", on_click=lambda e: do_delete_warehouse()),
        ],
        actions_alignment=ft.MainAxisAlignment.END,
    )

    dlg_create = ft.AlertDialog(
        modal=True,
        title=ft.Text("Crear un almacén"),
        shape=ft.RoundedRectangleBorder(radius=5),
        content=ft.Column([name_tf, color_dd], tight=True, spacing=10, width=400),
        actions=[
            ft.TextButton("Cancelar", on_click=lambda e: (setattr(dlg_create, "open", False), page.update(), close_dialog()),
                          style=ft.ButtonStyle(shape=ft.RoundedRectangleBorder(radius=5))),
            ft.FilledButton("Guardar", on_click=lambda e: save_warehouse(),
                            style=ft.ButtonStyle(shape=ft.RoundedRectangleBorder(radius=5))),
        ],
        actions_alignment=ft.MainAxisAlignment.END,
    )

    # Diálogo para elegir almacén al importar
    wh_dd = ft.Dropdown(label="Selecciona un almacén", width=360, on_change=on_pick_wh_changed)
    replace_stock_cb = ft.Checkbox(
        label="Reemplazar existencias en el almacén (no sumar)",
        value=False,
        on_change=lambda e: ui_state.__setitem__("replace_stock", bool(e.control.value)),
    )
    dlg_pick_wh = ft.AlertDialog(
        modal=True,
        title=ft.Text("¿A qué almacén se agregarán estos productos?"),
        content=ft.Column([wh_dd, replace_stock_cb], spacing=10, width=380, tight=True),
        actions=[ft.TextButton("Cancelar", on_click=on_pick_wh_cancel), ft.FilledButton("Confirmar", on_click=on_pick_wh_confirm)],
        actions_alignment=ft.MainAxisAlignment.END,
    )

    # Diálogo de ENTRADA DE PRODUCTOS
    dlg_entry = ft.AlertDialog(
        modal=True,
        title=ft.Text("Entrada de productos"),
        content=ft.Column(
            width=560, spacing=12,
            controls=[
                entry_wh_dd,
                entry_code_tf,
                ft.Divider(),
                ft.Text("Productos en esta entrada:", size=12, color=ft.Colors.GREY_700),
                entry_lines_col,
            ], height=400, scroll=ft.ScrollMode.AUTO
        ), shape=ft.RoundedRectangleBorder(radius=5),
        actions=[
            ft.TextButton("Vaciar", on_click=entry_clear),
            ft.TextButton("Cancelar", on_click=lambda e: (setattr(dlg_entry, "open", False), page.update(), close_dialog())),
            ft.FilledButton("Confirmar", on_click=entry_confirm),
        ],
        actions_alignment=ft.MainAxisAlignment.END,
    )

    alert_no_wh = ft.AlertDialog(
        modal=True,
        title=ft.Text("No hay almacenes"),
        content=ft.Text("Debes crear al menos un almacén antes de continuar."),
        actions=[ft.FilledButton("Crear un almacén", on_click=goto_create)],
        actions_alignment=ft.MainAxisAlignment.END,
    )

    def gradient_for(key: str):
        mp = {k:v for k,v in COLOR_CHOICES}; colors = mp.get(key, mp["slate"])
        return ft.LinearGradient(begin=ft.alignment.top_left, end=ft.alignment.bottom_right, colors=colors)

    def confirm_delete(w):
        warehouse_to_delete["id"] = w["id"]; warehouse_to_delete["name"] = w["name"]
        open_dialog(dlg_delete)

    def do_delete_warehouse():
        try:
            db.delete_warehouse_cascade(warehouse_to_delete["id"])
            notify("success", f"Almacén '{warehouse_to_delete['name']}' eliminado.")
        except Exception as ex:
            notify("error", f"Error: {ex}")
        finally:
            dlg_delete.open = False; page.update(); close_dialog(); render_warehouses()

    # Menubar
    menubar = ft.MenuBar(
        style=ft.MenuStyle(
            alignment=ft.alignment.top_left, bgcolor=ft.Colors.WHITE,
            mouse_cursor={ft.ControlState.HOVERED: ft.MouseCursor.WAIT, ft.ControlState.DEFAULT: ft.MouseCursor.ZOOM_OUT},
            shape=ft.RoundedRectangleBorder(radius=0),
        ),
        controls=[
            ft.SubmenuButton(
                content=ft.Text("Almacén"),
                on_open=handle_submenu_open, on_close=handle_submenu_close,
                controls=[
                    ft.MenuItemButton(content=ft.Text("Ver almacenes"), leading=ft.Icon(ft.Icons.INVENTORY),
                                    style=ft.ButtonStyle(bgcolor={ft.ControlState.HOVERED: ft.Colors.GREY_200},
                                                        shape=ft.RoundedRectangleBorder(radius=0)),
                                    on_click=handle_menu_item_click),
                    ft.MenuItemButton(content=ft.Text("Crear un almacen"), leading=ft.Icon(ft.Icons.WAREHOUSE),
                                    style=ft.ButtonStyle(bgcolor={ft.ControlState.HOVERED: ft.Colors.GREY_200},
                                                        shape=ft.RoundedRectangleBorder(radius=0)),
                                    on_click=handle_menu_item_click),
                    ft.MenuItemButton(content=ft.Text("Entrada de productos"), leading=ft.Icon(ft.Icons.QR_CODE),
                                    style=ft.ButtonStyle(bgcolor={ft.ControlState.HOVERED: ft.Colors.GREY_200},
                                                        shape=ft.RoundedRectangleBorder(radius=0)),
                                    on_click=handle_menu_item_click),
                    ft.MenuItemButton(content=ft.Text("Salida de productos"), leading=ft.Icon(ft.Icons.EXIT_TO_APP),
                                style=ft.ButtonStyle(bgcolor={ft.ControlState.HOVERED: ft.Colors.GREY_200},
                                                    shape=ft.RoundedRectangleBorder(radius=0)),
                                on_click=open_exit_dialog),
                ],
            ),

            ft.SubmenuButton(
                content=ft.Text("Productos"),
                on_open=handle_submenu_open, on_close=handle_submenu_close,
                controls=[
                    ft.MenuItemButton(content=ft.Text("Ver productos"), data="products_view",
                                    leading=ft.Icon(ft.Icons.INVENTORY_SHARP),
                                    style=ft.ButtonStyle(bgcolor={ft.ControlState.HOVERED: ft.Colors.GREY_200},
                                                        shape=ft.RoundedRectangleBorder(radius=0)),
                                    on_click=handle_menu_item_click),
                    ft.MenuItemButton(content=ft.Text("Agregar Lista de Productos"), data="products_import",
                                    leading=ft.Icon(ft.Icons.FILE_OPEN),
                                    style=ft.ButtonStyle(bgcolor={ft.ControlState.HOVERED: ft.Colors.GREY_200},
                                                        shape=ft.RoundedRectangleBorder(radius=0)),
                                    on_click=handle_menu_item_click),
                ],
            ),
            ft.SubmenuButton(
                content=ft.Text("Buscar"),
                on_open=handle_submenu_open, on_close=handle_submenu_close,
                controls=[
                    ft.MenuItemButton(content=ft.Text("Buscar un producto"),
                                    leading=ft.Icon(ft.Icons.SEARCH),
                                    style=ft.ButtonStyle(bgcolor={ft.ControlState.HOVERED: ft.Colors.GREY_200},
                                                        shape=ft.RoundedRectangleBorder(radius=0)),
                                    on_click=handle_menu_item_click),
                ],
            ),
        ],
    )


    top_bar = ft.Container(content=menubar, padding=0, left=0, right=0, top=0)

    page.add(
        ft.SafeArea(
            content=ft.Stack(controls=[top_bar, content_area]),
            top=True, bottom=True, left=True, right=True, expand=True,
        )
    )
    render_warehouses()


if __name__ == "__main__":
    ft.app(target=main)
